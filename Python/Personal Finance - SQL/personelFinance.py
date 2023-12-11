import sqlite3
import datetime
import openpyxl

db = sqlite3.connect('Expenses.sqlite')
cursor = db.cursor()

def main():

    answer = input('What do you want to do?\n 1- Add expense\n 2- Export expense to excel\n 3- Exit\n Please enter a number:')

    if answer == '1':
        script = 'SELECT * FROM Category'
        cursor.execute(script)
        rows = cursor.fetchall()
        for row in rows:
            print(row)
        Expense.from_input()
        print('---You succesfully entered a new expense.---')
        main()

    elif answer == '2':
        export_to_excel()
        print('---Expenses table is exported to excel---')
        main()

    else:
        print('---You exited the program---')




class Expense:
    def __init__(self, amount, categoryId, date):
        self.amount = amount
        self.categoryId = categoryId
        self.date = date

        new_row = (amount, categoryId, date)
        script = 'INSERT INTO Expense (Amount, CategoryId, Date)  VALUES (?,?,?);'

        cursor.execute(script, new_row)
        db.commit()

        script = 'SELECT * FROM Expense'
        cursor.execute(script)

        rows = cursor.fetchall()

        for row in rows:
            print(row)



    @classmethod
    def from_input(cls):

        categoryId = input('To which category belongs this expense? ')
        amount = input('What is the amount of the expense? ')
        date = datetime.date.today()
        return cls(amount, categoryId, date)



def export_to_excel():
    filename = str(datetime.date.today()) + '.xlsx'

    wb = openpyxl.Workbook(filename)
    wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    first_header = ws['A1']
    second_header = ws['B1']
    third_header = ws['C1']

    first_header.value = 'Amount'
    second_header.value = 'Category'
    third_header.value = 'Date'


    script = 'SELECT Expense.Amount,  Category.Name, Expense.Date FROM Expense JOIN Category ON (Expense.CategoryId = Category.CategoryID) '

    cursor.execute(script)
    table = cursor.fetchall()

    for row in table:
        ws.append(row)

    wb.save(filename)
    wb.close()
    db.close()



main()




